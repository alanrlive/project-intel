import email
import json
import logging
import re
from datetime import date, datetime
from pathlib import Path

import pdfplumber
import PyPDF2
from docx import Document as DocxDocument
from sqlalchemy.orm import Session

from app.llm_service import OllamaService
from app.models import Action, Deadline, Dependency, Document, DocumentType, RaidItemHistory, Risk, ScopeItem
from app.vector_service import VectorService
from app.config import get_llm_logging, get_model_assignments, LOGS_DIR

logger = logging.getLogger(__name__)


def _log_llm_response(
    file_path: Path,
    doc_type_name: str,
    prompt: str,
    raw_response: dict,
    parsed_ok: bool,
) -> None:
    """Write a JSON log file for the LLM extraction call. Never raises."""
    try:
        if not get_llm_logging():
            return
        now = datetime.now()
        log_dir = LOGS_DIR / "llm" / now.strftime("%Y-%m-%d")
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file = log_dir / f"{file_path.stem}_{now.strftime('%H%M%S')}.json"
        model = get_model_assignments().get("extraction", {}).get("model", "unknown")
        payload = {
            "document":      file_path.name,
            "document_type": doc_type_name,
            "model":         model,
            "timestamp":     now.isoformat(),
            "prompt_sent":   prompt,
            "raw_response":  json.dumps(raw_response),
            "parsed_ok":     parsed_ok,
        }
        log_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception:
        logger.warning("Failed to write LLM response log")

# ── Legacy extraction prompt (fallback only) ──────────────────────────────────
# Used by: run_llm_extraction() → process_document() → POST /documents/upload
# New uploads should use extract_with_type() which reads the prompt from the
# DocumentType table.  This constant is kept so the old single-file endpoint
# keeps working without a document_type_id.

EXTRACTION_PROMPT = """Analyze this {doc_type} document and extract structured information.

Return ONLY valid JSON with these exact keys (use empty arrays if nothing found):
{{
  "actions": [
    {{
      "description": "string",
      "owner": "string or null",
      "due_date": "YYYY-MM-DD or null",
      "priority": "high|medium|low",
      "status": "open|done|cancelled"
    }}
  ],
  "deadlines": [
    {{
      "description": "string",
      "deadline_date": "YYYY-MM-DD",
      "met": true or false
    }}
  ],
  "risks": [
    {{
      "description": "string",
      "impact": "high|medium|low",
      "likelihood": "high|medium|low",
      "mitigation": "string or null",
      "status": "open|closed"
    }}
  ],
  "dependencies": [
    {{
      "task_a": "string",
      "task_b": "string",
      "dependency_type": "blocks|enables|relates_to"
    }}
  ],
  "scope_items": [
    {{
      "description": "string",
      "source": "original_plan|change_request|meeting|deferred"
    }}
  ]
}}

EXTRACTION RULES:
- STATUS: Actions marked COMPLETED/DONE/FINISHED → "done". CANCELLED/DEFERRED → "cancelled". Otherwise → "open".
- STATUS: RISKS: Extract ALL risks including resolved/past ones from ANY section:   Look for: "Risk:", risk descriptions with impact/likelihood mentioned,  sections titled "Risks", "Resolved Risks", "Risks That Cleared Up". RESOLVED/CLOSED/MITIGATED/COMPLETELY GONE/NO LONGER AN ISSUE → status="closed" Otherwise → status="open". Example: Extract narrative risks: "X is no longer a problem", "That risk is gone"
- DEADLINES: Marked MET/ACHIEVED/DELIVERED → met=true. Otherwise → met=false.
- DATES: Extract ONLY from task text ("Due March 15", "by Feb 20", "deadline: 2025-03-10").  NEVER use the document's creation date, last modified date, or header/footer dates. If year missing, assume current year ({current_year}). Convert all dates to YYYY-MM-DD format for output. If no date in task description, return null.
- DEPENDENCIES: "A blocks B" means A must finish before B starts.
- SCOPE: Include future features, ideas marked DEFERRED/V3/OUT OF SCOPE.
- IMPORTANT: Extract ALL risks including resolved ones. Past risks are valuable historical data.

Document content:
{content}"""


# ── Text extractors ───────────────────────────────────────────────────────────

def _extract_pdf(file_path: Path) -> str:
    """Try pdfplumber first (better layout), fall back to PyPDF2."""
    try:
        with pdfplumber.open(file_path) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
            text = "\n".join(pages).strip()
            if text:
                return text
    except Exception as exc:
        logger.warning("pdfplumber failed for %s: %s — trying PyPDF2", file_path.name, exc)

    # PyPDF2 fallback
    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        return "\n".join(
            page.extract_text() or "" for page in reader.pages
        ).strip()


def _extract_docx(file_path: Path) -> str:
    doc = DocxDocument(str(file_path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_eml(file_path: Path) -> str:
    with open(file_path, "rb") as f:
        msg = email.message_from_binary_file(f)

    parts = []
    # Subject + sender as context
    parts.append(f"From: {msg.get('From', '')}")
    parts.append(f"To: {msg.get('To', '')}")
    parts.append(f"Subject: {msg.get('Subject', '')}")
    parts.append(f"Date: {msg.get('Date', '')}")
    parts.append("")

    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    parts.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            parts.append(payload.decode(msg.get_content_charset() or "utf-8", errors="replace"))

    return "\n".join(parts)


def _sheet_to_markdown(ws) -> str:
    """Convert a single openpyxl worksheet to a Markdown table. Returns empty string if blank."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        rows.append([str(cell) if cell is not None else "" for cell in row])

    if not rows:
        return ""

    header = rows[0]
    data_rows = rows[1:]

    col_widths = [max(len(h), 3) for h in header]
    for row in data_rows:
        for i, cell in enumerate(row):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(cell))

    def fmt_row(cells: list[str]) -> str:
        padded = [
            cells[i].ljust(col_widths[i]) if i < len(col_widths) else cells[i]
            for i in range(len(col_widths))
        ]
        return "| " + " | ".join(padded) + " |"

    separator = "| " + " | ".join("-" * w for w in col_widths) + " |"
    lines = [fmt_row(header), separator]
    for row in data_rows:
        padded = row + [""] * (len(col_widths) - len(row))
        lines.append(fmt_row(padded))

    return "\n".join(lines)


def _excel_to_markdown(file_path: Path) -> str:
    """
    Convert all sheets of an Excel workbook to Markdown tables separated by sheet headings.
    Ignores formatting/colours/formulas — values only.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        table = _sheet_to_markdown(ws)
        if table:
            sections.append(f"## {sheet_name}\n\n{table}")

    return "\n\n".join(sections)


def extract_text(file_path: Path) -> str:
    """Dispatch to the correct extractor based on file suffix."""
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf(file_path)
    if suffix == ".docx":
        return _extract_docx(file_path)
    if suffix in {".txt", ".md", ".markdown"}:
        return file_path.read_text(encoding="utf-8", errors="replace")
    if suffix in {".eml", ".msg"}:
        return _extract_eml(file_path)
    if suffix in {".xlsx", ".xls"}:
        return _excel_to_markdown(file_path)
    # Attempt plain text for unknown types
    logger.warning("Unknown file type %s — attempting plain text read", suffix)
    return file_path.read_text(encoding="utf-8", errors="replace")


# ── LLM extraction ────────────────────────────────────────────────────────────

async def run_llm_extraction(content: str, doc_type: str, llm: OllamaService) -> dict:
    """Send document content to Ollama and parse structured JSON response."""
    # Truncate very large documents to avoid overwhelming the context window
    max_chars = 12_000
    if len(content) > max_chars:
        logger.warning("Document truncated from %d to %d chars for extraction", len(content), max_chars)
        content = content[:max_chars] + "\n\n[... document truncated ...]"

    prompt = EXTRACTION_PROMPT.format(doc_type=doc_type, content=content)
    return await llm.extract(prompt)


# ── Open RAID item context ────────────────────────────────────────────────────

_OPEN_ITEM_CAP = 200


def _build_existing_items_context(db: Session) -> str:
    """
    Fetch all non-closed RAID items that have a reference_id and format them
    as a compact reconciliation block to inject into the extraction prompt.
    Returns an empty string if there are no qualifying items.
    """
    lines: list[str] = []

    queries: list[tuple[str, type]] = [
        ("ACT", Action),
        ("RSK", Risk),
        ("DL",  Deadline),
        ("DEP", Dependency),
        ("SCP", ScopeItem),
    ]

    for _prefix, model in queries:
        # Skip items without a reference_id — no reconciliation possible
        base = db.query(model).filter(model.reference_id.isnot(None))

        # Filter to non-closed/non-done rows where the model has a status column
        if hasattr(model, "status"):
            base = base.filter(model.status.notin_(["done", "closed", "cancelled"]))

        rows = base.limit(_OPEN_ITEM_CAP).all()
        for row in rows:
            if len(lines) >= _OPEN_ITEM_CAP:
                break
            desc = row.description.replace("\n", " ") if row.description else ""
            status = getattr(row, "status", "open") or "open"
            lines.append(f"{row.reference_id} | {desc} | {status}")

    if not lines:
        return ""

    items_block = "\n".join(lines)
    return (
        "\n\nExisting open items:\n"
        f"{items_block}\n\n"
        "Where a document references an existing item ID, return that ID in the id field "
        "so the system can reconcile it. New items with no matching ID should have id: null."
    )


# ── Universal extraction prompt ───────────────────────────────────────────────
# Runs for every upload regardless of document type.
# Document type is injected as a hint to bias focus, not to restrict scope.

UNIVERSAL_BASE_PROMPT = (
    'Extract all RAID items from this document. Return JSON (omit empty arrays):\n'
    '\n'
    '- "risks": [{"id": str|null, "description": str, "impact": "high"|"medium"|"low", "likelihood": "high"|"medium"|"low", "mitigation": str|null, "status": "open"|"closed"}]\n'
    '- "actions": [{"id": str|null, "description": str, "owner": str|null, "due_date": "YYYY-MM-DD"|null, "priority": "high"|"medium"|"low", "status": "open"|"done"|"cancelled"}]\n'
    '- "dependencies": [{"id": str|null, "task_a": str, "dependency_type": "blocks"|"enables"|"relates_to", "task_b": str, "notes": str|null}]\n'
    '- "deadlines": [{"id": str|null, "description": str, "deadline_date": "YYYY-MM-DD"|null, "met": bool}]\n'
    '- "scope_items": [{"id": str|null, "description": str, "source": str|null, "approved": bool, "impact_assessment": str|null}]\n'
    '\n'
    'STATUS: Actions COMPLETED/DONE/FINISHED/CLOSED="done". CANCELLED/DEFERRED="cancelled". Otherwise="open". '
    'Risks RESOLVED/CLOSED/MITIGATED="closed".\n'
    '\n'
    'CRITICAL: Return ALL items regardless of status. Do NOT omit completed or closed items. Return them with their correct status.\n'
    '\n'
    'FIELDS: Always return all fields for every item. Never truncate descriptions. '
    'For completed items carry forward original dates and priorities if present in the document.\n'
    '\n'
    'DESCRIPTION: Set description to the text as it appears in THIS document. '
    'For updates use the latest text not the original wording.\n'
    '\n'
    'IDs: Extract reference IDs exactly as they appear (e.g. ACT-001, RSK-003). Return null if no ID present.\n'
    '\n'
    'DATES: Use task dates not file metadata dates. Do NOT create deadline entries for action due dates — '
    'deadlines are distinct project milestones, not action completion dates.\n'
    '\n'
    'DEPENDENCIES: "A blocks B" means A must finish before B starts. '
    'Only extract explicit dependency relationships, not implied ones.\n'
    '\n'
    'JSON only, no markdown.'
)


# ── Type-aware extraction ─────────────────────────────────────────────────────

async def extract_with_type(
    file_path: Path,
    document_type_id: int,
    db: Session,
    llm: OllamaService,
) -> dict:
    """
    Extract all RAID items using the universal base prompt.
    Document type is injected as a hint to bias extraction focus.
    Custom document types may append an additional hint if extraction_prompt is set.
    """
    doc_type: DocumentType | None = db.query(DocumentType).filter(
        DocumentType.id == document_type_id
    ).first()

    type_name = doc_type.name if doc_type else "General"

    content = extract_text(file_path)
    if not content.strip():
        return {}

    # Persist raw content so vector embedding has something to work with
    doc = db.query(Document).filter(Document.file_path == str(file_path)).first()
    if doc and not doc.content_text:
        doc.content_text = content
        db.commit()

    # Truncate to avoid overwhelming Ollama context window
    max_chars = 12_000
    if len(content) > max_chars:
        logger.warning(
            "Document truncated from %d to %d chars for type '%s'",
            len(content), max_chars, type_name,
        )
        content = content[:max_chars] + "\n\n[... document truncated ...]"

    # Document type hint — biases focus without restricting scope
    type_hint = (
        f"Document type: {type_name}. Use this as context to guide extraction focus "
        f"but extract ALL relevant RAID items found regardless of type.\n\n"
    )

    # Custom types may carry an additional prompt hint (system types have none)
    custom_hint = ""
    if doc_type and doc_type.extraction_prompt:
        custom_hint = f"\nAdditional context for this document type: {doc_type.extraction_prompt}\n"

    existing_context = _build_existing_items_context(db)
    prompt = (
        f"{type_hint}"
        f"{UNIVERSAL_BASE_PROMPT}"
        f"{custom_hint}"
        f"{existing_context}"
        f"\n\nDocument content:\n{content}"
    )

    raw_response = await llm.extract(prompt)
    _log_llm_response(file_path, type_name, prompt, raw_response, bool(raw_response))
    return raw_response


# ── Date validation ───────────────────────────────────────────────────────────

def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return None


# ── DB storage ────────────────────────────────────────────────────────────────

def _write_history(
    db: Session,
    item_type: str,
    item_id: int,
    reference_id: str | None,
    description: str,
    status: str | None,
    source_document_id: int | None,
) -> None:
    db.add(RaidItemHistory(
        item_type=item_type,
        item_id=item_id,
        reference_id=reference_id,
        description=description,
        status=status,
        source_document_id=source_document_id,
    ))


# Matches any reference-id token at a word boundary (e.g. ACT-001, RSK-12, DL-003)
_REF_ID_RE = re.compile(r'\b[A-Z]{2,6}-\d+\b')


def _extract_source_text(content_text: str, reference_id: str) -> str | None:
    """
    Return the line(s) from content_text that mention reference_id.

    Algorithm:
    - Split on newlines; find the first line containing reference_id
      (case-insensitive).
    - If the next line exists and does NOT begin a new reference-id entry,
      append it (captures multi-line descriptions).
    - Strip whitespace; cap at 500 characters.
    - Returns None if reference_id is not found.
    """
    if not content_text or not reference_id:
        return None

    needle = reference_id.upper()
    lines  = content_text.splitlines()

    for i, line in enumerate(lines):
        if needle not in line.upper():
            continue
        text = line.strip()
        if i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            # Only merge if the next line doesn't start its own ref-id entry
            if next_line and not _REF_ID_RE.search(next_line):
                text = text + " " + next_line
        return text[:500] or None

    return None


def store_extracted_data(extracted: dict, doc_id: int, db: Session) -> dict:
    """
    Persist extracted actions/deadlines/risks/dependencies/scope_items.

    Reconciliation by reference_id:
    - reference_id present + existing row found → update changed fields, write history
    - reference_id present + no existing row  → insert new, write history
    - reference_id absent                      → insert new, write history

    Always writes a raid_item_history row on insert or update.
    Returns counts of rows inserted or updated.
    """
    counts: dict[str, int] = {
        "actions": 0,
        "deadlines": 0,
        "risks": 0,
        "dependencies": 0,
        "scope_items": 0,
    }

    # Fetch document content once — used to extract source text for history rows
    _doc = db.query(Document).filter(Document.id == doc_id).first()
    content_text: str = (_doc.content_text or "") if _doc else ""

    _valid_action_status = {"open", "done", "cancelled"}
    _valid_risk_status   = {"open", "closed"}

    # ── Actions ───────────────────────────────────────────────────────────────
    for item in extracted.get("actions", []):
        if not item.get("description"):
            continue
        ref_id     = item.get("id") or None
        raw_status = item.get("status", "open")
        status     = raw_status if raw_status in _valid_action_status else "open"

        existing: Action | None = (
            db.query(Action).filter(Action.reference_id == ref_id).first()
            if ref_id else None
        )

        if existing:
            incoming_desc = item["description"]
            if len(incoming_desc) >= len(existing.description):
                existing.description = incoming_desc
            existing.status   = status
            existing.owner    = item.get("owner") or existing.owner
            existing.due_date = _parse_date(item.get("due_date")) or existing.due_date
            existing.priority = item.get("priority") or existing.priority
            row_id    = existing.id
            hist_desc = (_extract_source_text(content_text, ref_id) if ref_id else None) or item["description"]
        else:
            row = Action(
                description=item["description"],
                owner=item.get("owner"),
                due_date=_parse_date(item.get("due_date")),
                priority=item.get("priority", "medium"),
                status=status,
                reference_id=ref_id,
                created_from_doc_id=doc_id,
            )
            db.add(row)
            db.flush()  # populate row.id
            row_id    = row.id
            hist_desc = item["description"]

        _write_history(db, "action", row_id, ref_id, hist_desc, status, doc_id)
        counts["actions"] += 1

    # ── Deadlines ─────────────────────────────────────────────────────────────
    for item in extracted.get("deadlines", []):
        dl = _parse_date(item.get("deadline_date"))
        if not item.get("description") or not dl:
            continue
        ref_id = item.get("id") or None
        met    = bool(item.get("met", False))

        existing: Deadline | None = (
            db.query(Deadline).filter(Deadline.reference_id == ref_id).first()
            if ref_id else None
        )

        if existing:
            incoming_desc = item["description"]
            if len(incoming_desc) >= len(existing.description):
                existing.description = incoming_desc
            existing.deadline_date = dl
            existing.met           = met
            row_id    = existing.id
            hist_desc = (_extract_source_text(content_text, ref_id) if ref_id else None) or item["description"]
        else:
            row = Deadline(
                description=item["description"],
                deadline_date=dl,
                met=met,
                reference_id=ref_id,
                source_doc_id=doc_id,
            )
            db.add(row)
            db.flush()
            row_id    = row.id
            hist_desc = item["description"]

        _write_history(db, "deadline", row_id, ref_id, hist_desc, "met" if met else "open", doc_id)
        counts["deadlines"] += 1

    # ── Risks ─────────────────────────────────────────────────────────────────
    for item in extracted.get("risks", []):
        if not item.get("description"):
            continue
        ref_id     = item.get("id") or None
        raw_status = item.get("status", "open")
        status     = raw_status if raw_status in _valid_risk_status else "open"

        existing: Risk | None = (
            db.query(Risk).filter(Risk.reference_id == ref_id).first()
            if ref_id else None
        )

        if existing:
            incoming_desc = item["description"]
            if len(incoming_desc) >= len(existing.description):
                existing.description = incoming_desc
            existing.status     = status
            existing.impact     = item.get("impact") or existing.impact
            existing.likelihood = item.get("likelihood") or existing.likelihood
            existing.mitigation = item.get("mitigation") or existing.mitigation
            row_id    = existing.id
            hist_desc = (_extract_source_text(content_text, ref_id) if ref_id else None) or item["description"]
        else:
            row = Risk(
                description=item["description"],
                impact=item.get("impact", "medium"),
                likelihood=item.get("likelihood", "medium"),
                mitigation=item.get("mitigation"),
                status=status,
                reference_id=ref_id,
            )
            db.add(row)
            db.flush()
            row_id    = row.id
            hist_desc = item["description"]

        _write_history(db, "risk", row_id, ref_id, hist_desc, status, doc_id)
        counts["risks"] += 1

    # ── Dependencies ──────────────────────────────────────────────────────────
    for item in extracted.get("dependencies", []):
        if not item.get("task_a") or not item.get("task_b"):
            continue
        ref_id   = item.get("id") or None
        dep_type = item.get("dependency_type", "relates_to")
        notes    = item.get("notes") or None

        existing: Dependency | None = None
        if ref_id:
            existing = db.query(Dependency).filter(Dependency.reference_id == ref_id).first()
        else:
            # No reference_id — deduplicate by (task_a, dependency_type, task_b)
            existing = (
                db.query(Dependency)
                .filter(
                    Dependency.task_a == item["task_a"],
                    Dependency.dependency_type == dep_type,
                    Dependency.task_b == item["task_b"],
                )
                .first()
            )

        if existing:
            existing.task_a          = item["task_a"]
            existing.task_b          = item["task_b"]
            existing.dependency_type = dep_type
            existing.notes           = notes or existing.notes
            row_id    = existing.id
            hist_desc = (_extract_source_text(content_text, ref_id) if ref_id else None) or f"{item['task_a']} {dep_type} {item['task_b']}"
        else:
            row = Dependency(
                task_a=item["task_a"],
                task_b=item["task_b"],
                dependency_type=dep_type,
                notes=notes,
                reference_id=ref_id,
            )
            db.add(row)
            db.flush()
            row_id    = row.id
            hist_desc = f"{item['task_a']} {dep_type} {item['task_b']}"

        _write_history(db, "dependency", row_id, ref_id, hist_desc, None, doc_id)
        counts["dependencies"] += 1

    # ── Scope items ───────────────────────────────────────────────────────────
    for item in extracted.get("scope_items", []):
        if not item.get("description"):
            continue
        ref_id = item.get("id") or None

        existing: ScopeItem | None = (
            db.query(ScopeItem).filter(ScopeItem.reference_id == ref_id).first()
            if ref_id else None
        )

        if existing:
            incoming_desc = item["description"]
            if len(incoming_desc) >= len(existing.description):
                existing.description = incoming_desc
            existing.source = item.get("source") or existing.source
            row_id    = existing.id
            hist_desc = (_extract_source_text(content_text, ref_id) if ref_id else None) or item["description"]
        else:
            row = ScopeItem(
                description=item["description"],
                source=item.get("source", "meeting"),
                approved=False,
                reference_id=ref_id,
            )
            db.add(row)
            db.flush()
            row_id    = row.id
            hist_desc = item["description"]

        _write_history(db, "scope_item", row_id, ref_id, hist_desc, None, doc_id)
        counts["scope_items"] += 1

    db.commit()

    # Embed document content into vector store for semantic search
    document = db.query(Document).filter(Document.id == doc_id).first()
    if document and document.content_text:
        vector_service = VectorService(db_path=Path("backend/data"))
        metadata = {
            "filename": document.filename,
            "upload_date": document.upload_date.isoformat() if document.upload_date else "",
            "document_type_id": document.document_type_id or 0,
        }
        success = vector_service.embed_document(document.id, document.content_text, metadata)
        if not success:
            logger.warning(
                "Failed to embed document %d, vector search unavailable for this doc",
                document.id,
            )

    return counts


# ── Top-level pipeline ────────────────────────────────────────────────────────

async def process_document(
    file_path: Path,
    original_filename: str,
    doc_type: str,
    db: Session,
    llm: OllamaService,
) -> dict:
    """
    Full pipeline: extract text → LLM extraction → store → return summary.
    The Document row must already exist in the DB (created by the router).
    Returns a summary dict with counts and any warnings.
    """
    # Fetch the document record (router already inserted it)
    doc: Document | None = db.query(Document).filter(
        Document.file_path == str(file_path)
    ).first()

    if doc is None:
        raise ValueError(f"No document record found for {file_path}")

    # 1. Extract text
    content = extract_text(file_path)
    if not content.strip():
        return {
            "doc_id": doc.id,
            "filename": original_filename,
            "warning": "No text could be extracted from this file.",
            "counts": {},
        }

    # 2. Persist content on the document row
    doc.content_text = content
    db.commit()

    # 3. LLM extraction
    try:
        extracted = await run_llm_extraction(content, doc_type, llm)
    except ValueError as exc:
        logger.error("LLM extraction failed for %s: %s", original_filename, exc)
        return {
            "doc_id": doc.id,
            "filename": original_filename,
            "warning": f"LLM returned malformed output: {exc}",
            "counts": {},
        }

    # 4. Store in DB
    counts = store_extracted_data(extracted, doc.id, db)

    return {
        "doc_id": doc.id,
        "filename": original_filename,
        "doc_type": doc_type,
        "counts": counts,
        "summary": _human_summary(counts),
    }


def _human_summary(counts: dict[str, int]) -> str:
    parts = []
    labels = {
        "actions":      ("action",     "actions"),
        "deadlines":    ("deadline",   "deadlines"),
        "risks":        ("risk",       "risks"),
        "dependencies": ("dependency", "dependencies"),
        "scope_items":  ("scope item", "scope items"),
    }
    for key, (singular, plural) in labels.items():
        n = counts.get(key, 0)
        if n:
            parts.append(f"{n} {singular if n == 1 else plural}")
    return f"Found {', '.join(parts)}." if parts else "No structured data extracted."
